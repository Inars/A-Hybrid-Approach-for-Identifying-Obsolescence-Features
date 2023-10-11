# Author: Elie SAAD
# Date: 11/08/2023
# Import the necessary modules
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from scipy.stats import pearsonr, gmean, norm
from sklearn.decomposition import PCA
from sklearn.tree import DecisionTreeClassifier, plot_tree
from sklearn.metrics import accuracy_score
import time
import pandas as pd


# Read the data from the Excel file
def read_excel_data(file_name, sheet_name):
    """
    Reads all data from the specified sheet in the Excel file.

    Args:
        file_name (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to read.

    Returns:
        list: A list of lists containing the data from the sheet.
    """

    workbook = openpyxl.load_workbook(file_name)  # Load the Excel file
    sheet = workbook[sheet_name]  # Get the sheet
    data = []  # Initialize the data list
    for row in sheet.rows:  # Iterate over the rows in the sheet
        data.append([cell.value for cell in row])  # Append the row to the data list
    return data  # Return the data list


# Remove rows based on last column value
def remove_rows(M, values_to_remove):
    """
    Removes all rows having the value of the last column as one of the values in the `values_to_remove` list.

    Args:
        M: A numpy matrix.
        values_to_remove: A list of values to remove from the last column of the matrix.

    Returns:
        A numpy matrix with the rows removed.
    """

    # Get the last column of the matrix.
    last_column = M[:, -1]

    # Get a boolean mask of the rows to remove.
    rows_to_remove_mask = np.isin(last_column, values_to_remove)

    # Remove the rows from the matrix.
    M = M[~rows_to_remove_mask]

    return M


# Convert the data to categorical
def to_categorical(X, column_numbers):
    """
    Changes the data within the matrix to categorical based on the column numbers provided.

    Args:
        X: The matrix of data to be converted to categorical.
        column_numbers: The column numbers to be converted to categorical.

    Returns:
        The matrix with the data converted to categorical as well as the dictionary of categories.
    """

    if X.ndim == 1:  # If input is a vector
        unique_vals = np.unique(X)  # Get the unique values
        category_dict = {
            val: idx for idx, val in enumerate(unique_vals)
        }  # Create a dictionary of categories
        return (
            np.array([category_dict[val] for val in X]),
            category_dict,
        )  # Return the vector with the data converted to categorical and the dictionary of categories

    elif X.ndim == 2:  # If input is a matrix
        for col in column_numbers:  # Iterate over the columns
            unique_vals = np.unique(X[:, col])  # Get the unique values
            category_dict = {
                val: idx for idx, val in enumerate(unique_vals)
            }  # Create a dictionary of categories
            X[:, col] = np.array(
                [category_dict[val] for val in X[:, col]]
            )  # Convert the column to categorical
        return (
            X,
            category_dict,
        )  # Return the matrix with the data converted to categorical and the dictionary of categories

    else:
        raise ValueError("Input must be a 1D or 2D array.")


# Mean of values
def mean_of_values(dictionary):
    """
    Returns the mean of the corresponding values within a dictionary within a dictionary.

    Args:
        dictionary: A dictionary where the values are lists.

    Returns:
        The mean of the values within the inner dictionary.
    """

    averages = {}  # Initialize the averages dictionary

    for (
        outer_key,
        inner_dict,
    ) in dictionary.items():  # Iterate through the outer dictionary
        total_values = {}  # Initialize the total values dictionary
        count_values = {}  # Initialize the count values dictionary

        for (
            inner_key,
            values,
        ) in inner_dict.items():  # Iterate through the inner dictionary
            for i, value in enumerate(values):  # Iterate through the values
                total_values[i] = (
                    total_values.get(i, 0) + value
                )  # Add the value to the total values
                count_values[i] = (
                    count_values.get(i, 0) + 1
                )  # Add 1 to the count values

        avg_values = [
            total_values[i] / count_values[i] for i in range(len(values))
        ]  # Calculate the average values
        averages[
            outer_key
        ] = avg_values  # Add the average values to the averages dictionary

    return averages  # Return the averages dictionary


# Count the number of repetitions of each element in a vector
def count_repeats(vector):
    """
    Returns the number of repetition of each element inside a vector.

    Args:
        vector (list): The vector to be counted.

    Returns:
        dict: A dictionary containing the number of repetition of each element.
    """

    counts = {}
    for element in vector:
        if element in counts:
            counts[element] += 1
        else:
            counts[element] = 1
    return counts


# Normalize boolean features
def normalize_boolean_features(data):
    """
    Normalizes the boolean features in a data matrix to a scale from 1 to 5.

    Args:
        data: The data matrix, where each row is a data point and each column is a feature.

    Returns:
        The normalized data matrix.
    """

    # boolean_feature_indices = np.where((data == 0) | (data == 1))[1] # Find the indices of the boolean features
    boolean_feature_indices = [0, 1, 5, 6, 7, 11]

    # Loop over data and scale boolean features
    scaled_data_matrix = data.copy()  # Initialize the scaled data matrix
    for idx in boolean_feature_indices:  # Iterate through the boolean feature indices
        boolean_feature = data[:, idx]  # Get the boolean feature

        std_dev = (
            np.std(boolean_feature) + 1e-10
        )  # Calculate the standard deviation and add a small value to the standard deviation to prevent it from ever being zero

        scaled_feature = (
            1 + 4 * (boolean_feature - np.mean(boolean_feature)) / std_dev
        )  # Scale boolean feature to a range of 1 to 5 based on normal distribution
        scaled_feature = np.clip(
            scaled_feature, 1, 5
        )  # Clip values to ensure they are within the range [1, 5]
        scaled_data_matrix[
            :, idx
        ] = scaled_feature  # Replace the original column with the scaled values

    return scaled_data_matrix  # Return the scaled data matrix


# Read the data from the Excel file
def read_data(file_name, sheet_name):
    """
    Reads the data from the Excel file.

    Args:
        file_name: The name of the Excel file.
        sheet_name: The name of the sheet in the Excel file.

    Returns:
        The data as a NumPy array.
    """

    # Read the data from the Excel file
    data = read_excel_data(file_name, sheet_name)  # Read the data from the Excel file
    data = np.array(data)  # Convert the data to a NumPy array
    data = data[
        :-6, 3:-2
    ]  # Remove the first 3 columns and the last 2 columns and the last 6 rows

    return data  # Return the data as a NumPy array


# Preprocess the data
def preprocess_data(data):
    """
    Preprocesses the data.

    Args:
        data: The data to be preprocessed of type NumPy array.

    Returns:
        The preprocessed data, the labels, the class names, the number of classes, and the feature names.
    """

    # Get the data
    feature_names = data[0, :-4]  # Get the feature names
    X = data[
        1:, :-4
    ]  # Get the instances and remove the dates column, the politique produit column, and the MTBF column
    y = data[1:, -1]  # Get the labels
    np.random.shuffle(X)  # Shuffle the instances

    # Convert the data to categorical
    column_numbers = [15]  # The column numbers to be converted to categorical

    X, _ = to_categorical(X, column_numbers)  # Convert the instances to categorical
    y, classes = to_categorical(y, [0])  # Convert the labels to categorical
    class_names = list(classes.keys())  # Get the class names
    number_of_classes = len(classes)  # Get the number of classes

    X = normalize_boolean_features(X)  # Normalize the boolean features

    return (
        X,
        y,
        class_names,
        number_of_classes,
        feature_names,
    )  # Return the preprocessed data, the labels, the class names, the number of classes and the feature names


# Pearson correlation coefficient
def pearson_correlation(data):
    """
    Calculates the Pearson correlation coefficient between all columns in the data matrix.

    Args:
        data: A NumPy array of real numbers.

    Returns:
        A NumPy array of correlation coefficients.
    """

    correlation = np.zeros(
        (data.shape[1], data.shape[1])
    )  # Initialize the correlation matrix
    for i in range(data.shape[1]):  # Iterate over the columns
        for j in range(data.shape[1]):  # Iterate over the columns
            if i != j:  # If the columns are different
                correlation[i, j] = pearsonr(data[:, i], data[:, j])[
                    0
                ]  # Calculate the correlation coefficient
    return correlation  # Return the correlation matrix


# Backward elimination
def backward_elimination(correlation, threshold):
    """
    Performs backward elimination on the correlation matrix to select the most important features.

    Args:
        correlation: A NumPy array of correlation coefficients.
        threshold: The threshold for feature selection.

    Returns:
        A list of the selected features.
    """

    selected_features = []  # Initialize the list of selected features
    for i in range(correlation.shape[0]):  # Iterate over the columns
        if len(selected_features) == 0:  # If no feature has been selected yet
            selected_features.append(i)  # Add the feature to the list
        else:  # If at least one feature has been selected
            max_correlation = np.max(
                correlation[selected_features, i]
            )  # Get the maximum correlation coefficient
            if (
                max_correlation < threshold
            ):  # If the maximum correlation coefficient is less than the threshold
                continue  # Skip the feature
            else:  # If the maximum correlation coefficient is greater than the threshold
                selected_features.append(i)  # Add the feature to the list
    return selected_features  # Return the list of selected features


# Principal component analysis
def principal_component_analysis(data, n_components, selected_features):
    """
    Calculates the principal components of the data.

    Args:
        data: A NumPy array of real numbers.
        n_components: The number of components to use in PCA.
        selected_features: The indices of the selected features obtained from backward elimination.

    Returns:
        The reduced data, the principal components, and the explained variances.
    """

    pca = PCA(n_components=n_components)  # PCA with n components
    pca.fit(data[:, selected_features])  # Fit the PCA model
    data_reduced = pca.transform(
        data[:, selected_features]
    )  # Transform the data to the new space
    principal_components = pca.components_  # Get the principal components
    explained_variances = pca.explained_variance_ratio_  # Get the explained variances
    return (
        data_reduced,
        principal_components,
        explained_variances,
        pca,
    )  # Return the reduced data, the principal components, and the explained variances


# Decision tree classifier
def decision_tree_classifier(X, y):
    """
    Creates a decision tree classifier and fits it to the training data.

    Args:
        X: The matrix of data.
        y: The vector of labels.

    Returns:
        The decision tree classifier.
    """

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42
    )  # Split the data into training and testing sets
    clf = DecisionTreeClassifier()  # Create a decision tree classifier
    clf.fit(X_train, y_train)  # Fit the classifier to the training set
    y_pred = clf.predict(X_test)  # Make predictions on the test set
    accuracy = accuracy_score(y_test, y_pred)  # Calculate the accuracy
    return accuracy, clf  # Return the accuracy and the classifier


# Find the best threshold
def find_threshold(X, y, n_components):
    """
    Finds the best threshold for feature selection.

    Args:
        X: The matrix of data.
        y: The vector of labels.
        n_components: The number of components to use in PCA.

    Returns:
        The best threshold.
    """

    # Find the best threshold
    best_accuracy = 0  # The best accuracy
    best_threshold = -1.00  # The best threshold
    for threshold in np.arange(-1.0, 1.01, 0.01):  # Iterate through the thresholds
        # Pearson correlation method
        correlation = pearson_correlation(
            X
        )  # Pearson correlation coefficient calculation
        selected_features = backward_elimination(
            correlation, threshold
        )  # Backward elimination of features

        # Check if the number of selected features is less than the number of components
        if (
            len(selected_features) < n_components
        ):  # If the number of selected features is less than the number of components
            n_components = len(
                selected_features
            )  # Set the number of components to the number of selected features

        # PCA
        X_reduced, _, _ = principal_component_analysis(
            X, n_components, selected_features
        )  # Transform the data to the new space

        # Decision tree classifier
        accuracy, _ = decision_tree_classifier(
            X_reduced, y
        )  # Create a decision tree classifier and fit it to the training data

        if accuracy > best_accuracy:  # If the accuracy is better than the best accuracy
            best_accuracy = accuracy  # Update the best accuracy
            best_threshold = threshold  # Update the best threshold

    threshold = best_threshold  # The threshold for feature selection

    return threshold


# Find the best number of components
def find_n_components(X, y, threshold):
    """
    Finds the best number of components to use in PCA.

    Args:
        X: The matrix of data.
        y: The vector of labels.
        threshold: The threshold for feature selection.

    Returns:
        The best number of components.
    """

    # Pearson correlation coefficient
    correlation = pearson_correlation(X)  # Pearson correlation coefficient calculation
    selected_features = backward_elimination(
        correlation, threshold
    )  # Backward elimination of features

    # Check if the number of selected features is less than the number of components
    number_of_features = X.shape[1]  # The number of selected features
    if (
        len(selected_features) < number_of_features
    ):  # If the number of selected features is less than the number of components
        n_components = len(
            selected_features
        )  # Set the number of components to the number of selected features
    else:  # If the number of selected features is greater than or equal to the number of components
        n_components = (
            number_of_features  # Set the number of components to the number of features
        )

    # Find the best number of components
    best_accuracy = 0  # The best accuracy
    best_n_components = 0.00  # The best threshold
    for i in np.arange(n_components, 0, -1):  # Iterate through the number of components
        # PCA
        X_reduced, _, _ = principal_component_analysis(
            X, i, selected_features
        )  # Transform the data to the new space

        # Decision tree classifier
        accuracy, _ = decision_tree_classifier(
            X_reduced, y
        )  # Create a decision tree classifier and fit it to the training data

        if accuracy > best_accuracy:  # If the accuracy is better than the best accuracy
            best_accuracy = accuracy  # Update the best accuracy
            best_n_components = i  # Update the best threshold

    n_components = best_n_components  # The number of components to use in PCA

    return n_components


# Identify the original features that contributed the most to each principal component
def identify_original_features(
    principal_components,
    feature_importances,
    explained_variances,
    print_individual_results,
    n_components,
):
    """
    Identifies the original features that contributed the most to each principal component.

    Args:
        principal_components: The principal components.
        feature_importances: The feature importances.
        explained_variances: The explained variances.
        print_individual_results: Whether to print the results of each test or not.
        result: The dictionary containing the results.
        n_components: The number of components to use in PCA.

    Returns:
        The dictionary containing the results.
    """

    # Identify the original features that contributed the most to each principal component
    result = {}  # Initialize the result dictionary
    for i in range(n_components):  # Iterate through the principal components
        component = principal_components[i]  # Get the current principal component
        abs_contributions = np.abs(
            component
        )  # Get the absolute values of the contributions
        result[i + 1] = [
            contribution * feature_importances[i] for contribution in abs_contributions
        ]  # Add the contributions to the results dictionary taking into account the feature importances

        # Print the top contributing feature indices and their names for each test
        if print_individual_results:  # If printing the results of each test is enabled
            top_contributions_indices = np.argsort(result[i + 1])[
                ::-1
            ]  # Sort and get the indices of the features with the highest contributions
            print(
                f"Principal Component {i + 1} - Explained Variance: {explained_variances[i]:.4f}"
            )  # Print the principal component and the explained variance
            for (
                idx
            ) in (
                top_contributions_indices
            ):  # Iterate through the top contributing feature indices
                print(
                    f"Feature {idx}: {result[i + 1][idx]}"
                )  # Print the feature index and its contribution
            print()  # Print a new line

    return result


# The entire algorithm
def algorithm(X, y, threshold, n_components):
    """
    Runs the algorithm.

    Args:
        X: The matrix of data.
        y: The vector of labels.
        threshold: The threshold for feature selection.
        n_components: The number of components to use in PCA.

    Returns:
        The accuracy, the feature importances, the principal components, the explained variances, the decision tree classifier, the reduced data, and the selected features.
    """

    # Pearson correlation coefficient
    correlation = pearson_correlation(X)  # Pearson correlation coefficient calculation
    selected_features = backward_elimination(
        correlation, threshold
    )  # Backward elimination of features

    # PCA
    (
        X_reduced,
        principal_components,
        explained_variances,
        pca,
    ) = principal_component_analysis(
        X, n_components, selected_features
    )  # Transform the data to the new space

    # Decision tree classifier
    accuracy, clf = decision_tree_classifier(
        X_reduced, y
    )  # Create a decision tree classifier and fit it to the training data
    feature_importances = (
        clf.feature_importances_
    )  # Calculate the importances of each feature

    return (
        accuracy,
        feature_importances,
        principal_components,
        explained_variances,
        clf,
        X_reduced,
        selected_features,
        pca,
    )  # Return the accuracy, the feature importances, the principal components, the explained variances, the decision tree classifier, the reduced data, and the selected features


# Plot the feature importances
def plot_feature_importances(X, feature_importances):
    """
    Plots the feature importances.

    Args:
        X: The matrix of data.
        feature_importances: The feature importances.
    """

    X_train, _, _, _ = train_test_split(
        X, y, test_size=0.2, random_state=42
    )  # Split the data into training and testing sets

    # Get feature names and sort indices
    feature_names = [
        f"Feature {i}" for i in range(X_train.shape[1])
    ]  # Get the feature names
    sorted_indices = np.argsort(feature_importances)[::-1]  # Get the sorted indices

    # Plot feature importances
    plt.figure(figsize=(10, 6))  # Create a figure
    plt.bar(
        range(X_train.shape[1]), feature_importances[sorted_indices]
    )  # Plot the feature importances
    plt.xticks(
        range(X_train.shape[1]),
        np.array(feature_names)[sorted_indices],
        rotation=45,
        ha="right",
    )  # Set the x-ticks
    plt.xlabel("Feature")  # Set the x-label
    plt.ylabel("Importance")  # Set the y-label
    plt.title("Feature Importance")  # Set the title
    plt.tight_layout()  # Set the layout
    plt.show()  # Show the plot


# Plot the decision tree
def plot_decision_tree(clf, X, class_names):
    """
    Plots the decision tree.

    Args:
        clf: The decision tree classifier.
        X_train: The training data.
        class_names: The names of the classes.
    """

    # Get feature names
    X_train, _, _, _ = train_test_split(
        X, y, test_size=0.2, random_state=42
    )  # Split the data into training and testing sets
    feature_names = [
        f"Feature {i}" for i in range(X_train.shape[1])
    ]  # Get the feature names

    # Plot the decision tree
    plt.figure(figsize=(20, 10))  # Create a figure
    plot_tree(
        clf, filled=True, feature_names=feature_names, class_names=class_names
    )  # Plot the decision tree
    plt.tight_layout()  # Set the layout
    plt.show()  # Show the plot


# Plot the feature contributions
def plot_feature_contributions(
    feature_contributions_average,
    feature_contributions_best,
    feature_contributions_new,
    number_of_features,
    feature_names,
):
    """
    Plots the feature contributions.

    Args:
        feature_contributions_average: The average feature contributions.
        feature_contributions_best: The feature contributions from the best classifier.
        feature_contributions_new: The feature contributions of the best classifier applied to new data.
        number_of_features: The number of features.
        feature_names: The names of the features.
    """
    # Plot feature importances
    plt.figure(figsize=(27, 21))  # Set the figure size
    plt.rcParams.update({"font.size": 70})  # Set the font size
    plt.rc("font", family="Times New Roman")  # Set the font

    colors = [
        (74, 144, 226),  # Red, Green, Blue
        (208, 2, 27),  # Red, Green, Blue
        (126, 211, 33),  # Red, Green, Blue
    ]  # Create a list of colors for the bars
    colors = [
        (r / 255, g / 255, b / 255) for r, g, b in colors
    ]  # Normalize the RGB values to the range [0, 1]
    feature_names = [
        f"{i+1}" for i in range(feature_names.shape[0])
    ]  # Get the feature names

    # Plot each set of feature contributions with a different color
    plt.semilogy(
        np.arange(number_of_features) - 0.2,
        feature_contributions_average,
        color=colors[0],
    )
    plt.semilogy(
        np.arange(number_of_features), feature_contributions_best, color=colors[1]
    )
    plt.semilogy(
        np.arange(number_of_features) + 0.2, feature_contributions_new, color=colors[2]
    )
    plt.bar(
        np.arange(number_of_features) - 0.2,
        feature_contributions_average,
        width=0.2,
        label="Average",
        color=colors[0],
    )
    plt.bar(
        np.arange(number_of_features),
        feature_contributions_best,
        width=0.2,
        label="Best",
        color=colors[1],
    )
    plt.bar(
        np.arange(number_of_features) + 0.2,
        feature_contributions_new,
        width=0.2,
        label="New Data",
        color=colors[2],
    )
    plt.xticks(
        range(number_of_features), np.array(feature_names), rotation=45, ha="right"
    )  # Set the x-ticks
    plt.xlabel("Features")  # Set the x-label
    plt.ylabel("Importance")  # Set the y-label
    plt.tight_layout()  # Set the layout
    plt.show()  # Show the plot


# Plot the accuracies
def plot_accuracies(accuracies):
    """
    Plots the accuracies.

    Args:
        accuracies: The accuracies.
    """

    tests = list(range(1, len(accuracies) + 1))  # Generating a list of test numbers

    plt.figure(figsize=(27, 21))  # Set the figure size
    plt.rcParams.update({"font.size": 70})  # Set the font size
    plt.rc("font", family="Times New Roman")  # Set the font

    plt.plot(tests, accuracies)  # Creating the plot
    geo_mean = gmean(accuracies)  # Calculating the geometric mean
    plt.axhline(
        y=geo_mean, color="r", linestyle="--", label="Geometric Mean"
    )  # Plotting the geometric mean
    plt.xlabel("Test #")  # Setting the x-label
    plt.ylabel("Accuracy")  # Setting the y-label
    plt.grid(True)  # Setting the grid
    plt.show()  # Showing the plot


def count_yellow_cells_per_column(file_path, sheet_name):
    """
    Counts all the yellow colored cells within each column of an Excel file and returns a dictionary where each key is the column title (first row of the Excel file contains all the column titles) and the value is the number of cells that are colored in yellow.

    Args:
        file_path: The path to the Excel file.

    Returns:
        A dictionary where each key is the column title and the value is the number of cells that are colored in yellow.
    """

    yellow_counts = {}  # Initialize an empty dictionary to store the counts for each column

    workbook = openpyxl.load_workbook(file_path)  # Load the Excel file
    sheet = workbook[sheet_name]  # Get the sheet

    for row_index, row in enumerate(sheet.iter_rows()):
        for cell_index, cell in enumerate(row):
            # Check if the cell has a yellow fill color (assuming RGB value of yellow)
            if cell.fill.start_color.rgb == 'FFFFFF00':
                column_title = sheet.cell(row=1, column=cell_index + 1).value
                if column_title not in yellow_counts:
                    yellow_counts[column_title] = 0
                yellow_counts[column_title] += 1

    sorted_yellow_counts = dict(sorted(yellow_counts.items(), key=lambda item: item[1], reverse=True))  # Sort the dictionary by value in descending order

    return sorted_yellow_counts

# Main function
def main(
    n,
    file_name,
    sheet_name,
    n_components,
    threshold,
    optimise,
    plot_figures,
    print_individual_results,
    print_mean_contributions,
    plot_accuracies_bool,
    plot_feature_contributions_bool,
    remove_rows_bool,
):
    # Load the data
    data = read_data(file_name, sheet_name)  # Read the data from the Excel file

    # # Remove rows based on the last column value
    # if remove_rows_bool: # If removing rows based on the last column value is enabled
    #     data = remove_rows(data, ["stock", "LBO", "redesign mineur"]) # Remove rows based on the last column value

    # Initialize the variables
    clfs = []  # Initialize the decision tree classifiers list
    pcas = []  # Initialize the principal component analyses list
    accuracies = []  # Initialize the accuracies list
    results = {}  # Initialize the results dictionary
    if optimise:  # If optimisation is enabled
        threshold_list = []  # Initialize the list of thresholds
        n_components_list = []  # Initialize the list of number of components

    # Run the algorithm n times
    for i in range(n):
        start_time = time.time()  # Start the timer
        print(f"Test #{i + 1}")  # Print the test number

        X, y, class_names, number_of_classes, feature_names = preprocess_data(
            data
        )  # Preprocess the data

        # Find the optimal threshold and the optimal number of components
        if optimise:  # If optimisation is enabled
            threshold = find_threshold(X, y, n_components)  # Find the optimal threshold
            n_components = find_n_components(
                X, y, threshold
            )  # Find the optimal number of components
            threshold_list.append(threshold)  # Add the threshold to the list
            n_components_list.append(
                n_components
            )  # Add the number of components to the list
            print(f"Optimal Threshold: {threshold:.2f}")  # Print the optimal threshold
            print(
                f"Optimal Number of Components: {n_components}"
            )  # Print the optimal number of components

        (
            accuracy,
            feature_importances,
            principal_components,
            explained_variances,
            clf,
            X_reduced,
            selected_features,
            pca,
        ) = algorithm(
            X, y, threshold, n_components
        )  # Run the algorithm
        clfs.append(clf)  # Add the classifier to the list
        pcas.append(pca)  # Add the PCA to the list
        accuracies.append(accuracy)  # Add the accuracy to the list
        print(f"Accuracy: {accuracy:.4f}")  # Print the accuracy
        end_time = time.time()  # End the timer
        elapsed_time = end_time - start_time  # Calculate the elapsed time
        print("Elapsed time:", elapsed_time)  # Print the elapsed time
        print()  # Print a new line

        # Identify the original features that contributed the most to each principal component
        results[f"Test #{i + 1}"] = identify_original_features(
            principal_components,
            feature_importances,
            explained_variances,
            print_individual_results,
            n_components,
        )  # Identify the original features that contributed the most to each principal component

        # Plot the feature importances and the decision tree
        if plot_figures:  # If plotting the figures is enabled
            plot_feature_importances(
                X_reduced, feature_importances
            )  # Plot the feature importances
            plot_decision_tree(clf, X_reduced, class_names)  # Plot the decision tree

    if not optimise:  # If optimisation is disabled
        # Get the mean feature contributions
        mean_results = mean_of_values(results)  # Get the mean of the results
        mean_feature_contributions = np.zeros(
            len(next(iter(mean_results.values())))
        )  # Initialize the mean feature contributions

        for key, values in mean_results.items():  # Iterate through the results
            for i in range(len(values)):  # Iterate through the feature contributions
                mean_feature_contributions[i] += values[
                    i
                ]  # Add the feature contribution to the mean feature contribution
        mean_feature_contributions = [
            value / n for value in mean_feature_contributions
        ]  # Divide the mean feature contribution by the number of tests

    # Print the results
    print()  # Print a new line
    print("Results:")  # Print the results
    print("---------")  # Print a line
    print(f"Number of Tests Conducted: {n}")  # Print the number of tests
    if optimise:  # If optimisation is enabled
        print(
            f"Test #{np.argmax(accuracies) + 1} has the highest accuracy: {np.max(accuracies):.4f}"
        )  # Print the test with the highest accuracy
        print(
            f"Optimal Threshold  for Test #{np.argmax(accuracies) + 1}: {threshold_list[np.argmax(accuracies)]:.2f}"
        )  # Print the optimal threshold
        print(
            f"Optimal Number of Components for Test #{np.argmax(accuracies) + 1}: {n_components_list[np.argmax(accuracies)]}"
        )  # Print the optimal number of components
    else:  # If optimisation is disabled
        print(f"Number of Components: {n_components}")  # Print the number of components
        print(f"Threshold: {threshold:.2f}")  # Print the threshold
        print(f"Mean Accuracy: {np.mean(accuracies):.4f}")  # Print the mean accuracy
        print(
            f"Geometric Mean of Accuracy: {gmean(accuracies):.4f}"
        )  # Print the geometric mean of the accuracy
        print(
            f"Standard Deviation of Accuracy: {np.std(accuracies):.4f}"
        )  # Print the standard deviation of the accuracy
        print(
            f"Minimum Accuracy: {np.min(accuracies):.4f}"
        )  # Print the minimum accuracy
        print(
            f"Maximum Accuracy: {np.max(accuracies):.4f}"
        )  # Print the maximum accuracy
        print()  # Print a new line

        if (
            print_mean_contributions
        ):  # If printing the mean feature contributions is enabled
            print("Mean Feature Contributions:")  # Print the mean feature contributions
            top_contributions_indices = np.argsort(mean_feature_contributions)[
                ::-1
            ]  # Sort and get the indices of the features with the highest mean contribution
            for (
                idx
            ) in (
                top_contributions_indices
            ):  # Iterate through the top contributing feature indicess
                print(
                    f"Feature {selected_features[idx]}: {mean_feature_contributions[idx]}"
                )  # Print the mean feature contribution
            print()  # Print a new line

        # Deduce the feature contributions
        feature_contributions = np.zeros(
            X.shape[1]
        )  # Initialize the feature contributions
        for i in range(len(selected_features)):  # Iterate through the selected features
            feature_contributions[selected_features[i]] = mean_feature_contributions[
                i
            ]  # Add the mean feature contribution to the feature contributions
        print("All Feature Contributions:")  # Print the feature contributions
        top_feature_contributions_indices = np.argsort(feature_contributions)[
            ::-1
        ]  # Sort and get the indices of the features with the highest contribution
        for idx in top_feature_contributions_indices:  # Iterate through the features
            print(
                f"Feature {idx+1}: {feature_contributions[idx]} ({feature_names[idx]})"
            )  # Print the feature contribution

        # Plot the feature contributions
        if (
            plot_feature_contributions_bool
        ):  # If plotting the feature contributions is enabled
            best_clf = clfs[np.argmax(accuracies)]  # Get the best classifier
            best_pca = pcas[np.argmax(accuracies)]  # Get the best PCA

            _, X_test, _, y_test = train_test_split(
                X_reduced, y, test_size=0.2, random_state=42
            )  # Split the data into training and testing sets
            y_pred = best_clf.predict(X_test)  # Make predictions on the test set
            feature_importances = (
                clf.feature_importances_
            )  # Calculate the importances of each feature

            result = {}  # Initialize the result dictionary
            result["Result"] = identify_original_features(
                principal_components,
                feature_importances,
                explained_variances,
                print_individual_results,
                n_components,
            )  # Identify the original features that contributed the most to each principal component
            mean_result_best = mean_of_values(result)  # Get the mean of the results

            feature_contributions_best = np.zeros(
                X.shape[1]
            )  # Initialize the feature contributions
            for i in range(
                len(selected_features)
            ):  # Iterate through the selected features
                feature_contributions_best[selected_features[i]] = mean_result_best[
                    "Result"
                ][
                    i
                ]  # Add the mean feature contribution to the feature contributions

            # Read the new data and preprocess it
            data_new = read_excel_data(
                file_name, "nonobso"
            )  # Read the data from the Excel file
            data_new = np.array(data_new)  # Convert the data to a NumPy array
            data_new = data_new[
                :, 3:-1
            ]  # Remove the first 3 columns and the last 2 columns and the last 6 rows
            X_new, y_new, _, _, _ = preprocess_data(data_new)  # Preprocess the data

            # PCA
            X_new_reduced = best_pca.transform(
                X_new[:, selected_features]
            )  # Transform the data to the new space
            principal_components = best_pca.components_  # Get the principal components
            explained_variances = (
                best_pca.explained_variance_ratio_
            )  # Get the explained variances

            # Decision tree classifier
            y_pred = best_clf.predict(X_new_reduced)  # Make predictions on the test set
            accuracy = accuracy_score(y_new, y_pred)  # Calculate the accuracy
            feature_importances = (
                clf.feature_importances_
            )  # Calculate the importances of each feature
            print()
            print(f"New Data Accuracy: {accuracy:.4f}")  # Print the accuracy

            # Identify the original features that contributed the most to each principal component
            result_new = {}  # Initialize the result dictionary
            result_new["Result"] = identify_original_features(
                principal_components,
                feature_importances,
                explained_variances,
                print_individual_results,
                n_components,
            )  # Identify the original features that contributed the most to each principal component
            mean_result_new = mean_of_values(result_new)  # Get the mean of the results

            feature_contributions_new = np.zeros(
                X.shape[1]
            )  # Initialize the feature contributions
            for i in range(
                len(selected_features)
            ):  # Iterate through the selected features
                feature_contributions_new[selected_features[i]] = mean_result_new[
                    "Result"
                ][
                    i
                ]  # Add the mean feature contribution to the feature contributions

            # Plot all the mean feature contributions
            plot_feature_contributions(
                feature_contributions,
                feature_contributions_best,
                feature_contributions_new,
                X.shape[1],
                feature_names,
            )  # Plot the feature contributions

        # Plot the accuracies if enabled
        if plot_accuracies_bool:  # If plotting the accuracies is enabled
            plot_accuracies(accuracies)  # Plot the accuracies


# If the script is run as the main script
if __name__ == "__main__":  # If the script is run as the main script
    # Initialize the parameters
    n = 1000  # The number of times to run the main function
    file_name = "data.xlsx"  # The name of the Excel file
    sheet_name = "obso"  # The name of the sheet in the Excel file
    n_components = 7  # The number of components to use in PCA
    threshold = 0.15  # The threshold for feature selection
    optimise = False  # Whether to optimise the threshold and the number of components
    plot_figures = False  # Whether to plot the figures
    plot_accuracies_bool = True  # Whether to plot the accuracies
    plot_feature_contributions_bool = True  # Whether to plot the feature contributions
    print_individual_results = False  # Whether to print the results of each test
    print_mean_contributions = False  # Whether to print the mean feature contributions
    remove_rows_bool = False  # Whether to remove rows based on the last column value

    # Run the main code
    main(
        n,
        file_name,
        sheet_name,
        n_components,
        threshold,
        optimise,
        plot_figures,
        print_individual_results,
        print_mean_contributions,
        plot_accuracies_bool,
        plot_feature_contributions_bool,
        remove_rows_bool,
    )  # Run the main function

    # # Count the number of yellow cells in each column
    # column_yellow_cell_counts = count_yellow_cells_per_column(file_name, sheet_name) # Count the number of yellow cells in each column
    
    # # Print the results to the console.
    # for column_title, yellow_cell_count in column_yellow_cell_counts.items():
    #     print(f'Column title: "{column_title}", Yellow cell count: {yellow_cell_count}')
